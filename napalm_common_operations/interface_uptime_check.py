from napalm import get_network_driver
import textfsm
import re
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

def interface_uptime_check(network_connection, device_type):
    '''
    will print out a list of interfaces and their uptime based on last seen input/output packet
    (whichever is more recent)
    :return:
    '''
    # regex compile
    int_match = re.compile('([0-9]{2}:)+')
    week_day_match = re.compile('([0-9]+[wy][0-9]+[dw])')
    vlan_loopback_tunnel_match = re.compile('([vV]lan|[Ll]oopback|[tT]unnel)')

    try:
        # get facts to get switch hostname
        network_facts = network_connection.get_facts()
        hostname = network_facts['hostname']

        # stores all the interface info
        interface_summary = {
            hostname: []
        }

        # we need to pull the interface names into a list
        interface_list = [key for key, value in network_connection.get_interfaces().items()]

        # now we can do an 'show interface' to get all the stats, we'll need the textFSM template
        if device_type == 'cisco_ios' or device_type == 'ios':
            template_file = 'textfsm_templates/show_interfaces.textfsm'
        elif device_type == 'cisco_nxos' or device_type == 'nxos_ssh':
            return False

        with open(template_file) as template:
            # set the template
            template = textfsm.TextFSM(template)


            # the commands to run
            cmd = 'show interface'
            # get output of the command
            output = network_connection.cli([cmd])
            # parse the output
            parsed_output = template.ParseText(output[cmd])

            # go through parsed output for each interface
            for i in parsed_output:
                # if its a vlan interface, loopback or tunnel interface move on
                if re.match(vlan_loopback_tunnel_match, i[0]):
                    continue

                # variables to reference later in easier form
                interface_name = i[0]
                last_input = i[15]
                last_output = i[16]

                # if never input/output add to a list
                if last_input.lower() == 'never' and last_output.lower() == 'never':
                    interface_summary[hostname].append(
                        {
                            'interface': interface_name,
                            'never_input_output': True,
                            'extended_no_input_output': False,
                            'recent_input_output': False,
                            'last_input': last_input,
                            'last_output': last_output
                        }
                    )
                # if we match the week/day format in either its automatically extended usage
                elif re.match(week_day_match, last_input) or re.match(week_day_match, last_output):
                    interface_summary[hostname].append(
                        {
                            'interface': interface_name,
                            'never_input_output': False,
                            'extended_no_input_output': True,
                            'recent_input_output': False,
                            'last_input': last_input,
                            'last_output': last_output
                        }
                    )
                # otherwise...it is recently used
                else:
                    interface_summary[hostname].append(
                    {
                        'interface': interface_name,
                        'never_input_output': False,
                        'extended_no_input_output': False,
                        'recent_input_output': True,
                        'last_input': last_input,
                        'last_output': last_output
                    }
                )

            # return the data
            return interface_summary

    except Exception as e:
        print(e)

def excel_workbook_creation(interface_summary):
    # extract the hostname
    hostname = list(interface_summary.keys())[0]
    # create a pandas dataframe from the dictionary
    dataframe = pd.DataFrame.from_dict(data=interface_summary[hostname])

    try:
        print(f'writing info for {hostname} to input_output_interfaces.xlsx')
        # write the dataframe to excel, use ExcelWriter to open in append mode
        if os.path.exists('input_output_interfaces.xlsx'):
            # print('file exists')
            with pd.ExcelWriter('input_output_interfaces.xlsx', mode='a') as writer:
                dataframe.to_excel(writer, sheet_name=hostname, index=False)
        else:
            # print('file does not exist')
            with pd.ExcelWriter('input_output_interfaces.xlsx', mode='w') as writer:
                dataframe.to_excel(writer, sheet_name=hostname, index=False)

        # create a table in the excel sheet
        int_workbook = openpyxl.load_workbook('input_output_interfaces.xlsx')
        # open the tab with the hostname of the switch
        int_worksheet = int_workbook[hostname]
        table_style = TableStyleInfo(name='TableStyleMedium2',
                                     showRowStripes=True)
        # the use of int_worksheet.max_row will return the max row number to use for the table
        table = Table(ref=f'$A1:$F{int_worksheet.max_row}',
                      displayName=f'{hostname}_table', tableStyleInfo=table_style)
        # add to the worksheet the table over the selected rows
        int_worksheet.add_table(table)
        # save the xlsx
        int_workbook.save('input_output_interfaces.xlsx')
        # close the file
        int_workbook.close()
    except Exception as e:
        print(e)




