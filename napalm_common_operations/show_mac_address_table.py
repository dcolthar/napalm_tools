import textfsm
from operator import itemgetter
import json
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def get_mac_address_table(network_connection):
    cmd = 'show mac address-table dynamic'
    cmd2 = 'show ip arp'
    results = network_connection.cli([cmd, cmd2])

    mac_template_file = 'textfsm_templates/show_mac_address.textfsm'
    arp_template_file = 'textfsm_templates/show_ip_arp.textfsm'

    arp_dict = {}

    # parse arp results
    with open(arp_template_file) as template:
        template = textfsm.TextFSM(template)
        arp_parsed = template.ParseText(results[cmd2])
        # only add to list if Vlan is in the interface name
        for arp_entry in arp_parsed:
            if 'vlan' in arp_entry[5].lower():
                mac = arp_entry[3]
                ip = arp_entry[1]
                vlan = arp_entry[5][4::]
                arp_dict[mac] = {
                    'mac': mac,
                    'ip': ip,
                    'vlan': vlan
                }

    with open(mac_template_file) as template:
        template = textfsm.TextFSM(template)

        parsed_results = sorted(template.ParseText(results[cmd]), key=itemgetter(3))

        facts = network_connection.get_facts()
        hostname = facts['hostname']

        # stores the total parsed output to return
        total = {}

        for i in parsed_results:
            mac = i[0]
            vlan = i[2]
            # if the mac address has an arp entry as well...
            if mac in arp_dict and str(vlan) == arp_dict[mac]['vlan']:
                data_results = {
                    'VLAN': vlan,
                    'MAC': mac,
                    'IP_ADDRESS': arp_dict[mac]['ip']
                }
            else:
                data_results = {
                    'VLAN': vlan,
                    'MAC': mac,
                    'IP_ADDRESS': ''
                }
            # add to total list, first see if it exists and append to it if so elsewise make a new inner list
            if not i[3] in total:
                total[i[3]] = [data_results]
            else:
                total[i[3]].append(data_results)

        total = {hostname: total}
        return total

def write_mac_addresses_to_excel(mac_address_table_info):

    columns = [
        'Interface',
        'MAC Address Learned',
        'VLAN',
        'IP Address'
    ]

    # some styling things
    bold = Font(bold=True)
    white_text_bold = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    medium_right_up_down_border = Border(left=Side(style='thin'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))
    medium_left_up_down_border = Border(left=Side(style='medium'),
                                        right=Side(style='thin'),
                                        top=Side(style='medium'),
                                        bottom=Side(style='medium'))
    medium_top_bottom_border = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='medium'),
                                      bottom=Side(style='medium'))
    # the color you can find by getting the RGB value and then converting to hex
    grey_fill = PatternFill(
        start_color='595959',
        end_color='595959',
        fill_type='solid'
    )

    for site, closets in mac_address_table_info.items():
        row = 1
        column = 1

        output_file = f'mac_tables/site_{site}.xlsx'

        # try to start fresh on the file each time we do work
        try:
            os.remove(output_file)
        except:
            pass

        mac_workbook = openpyxl.Workbook()
        # index page
        mac_worksheet = mac_workbook.active
        mac_worksheet.title = 'index'

        for closet, switches in closets.items():

            for outer in switches:
                for switch, interfaces in outer.items():
                    # reset row
                    row = 1

                    # name of the sheet to create
                    this_sheetname = f'{switch}'

                    # create the sheet if it doesn't exist
                    if this_sheetname not in mac_workbook.sheetnames:
                        mac_workbook.create_sheet(this_sheetname)

                    # set the current sheet
                    if mac_workbook.active.title != this_sheetname:
                        mac_worksheet = mac_workbook[this_sheetname]
                        #mac_workbook.active(this_sheetname)

                    # build the column headers
                    mac_worksheet.column_dimensions['A'].width = 25
                    mac_worksheet.column_dimensions['B'].width = 25
                    mac_worksheet.column_dimensions['C'].width = 25
                    mac_worksheet.column_dimensions['D'].width = 25

                    for column_header in columns:
                        cell = mac_worksheet.cell(row=row, column=column)
                        cell.value = column_header
                        cell.fill = grey_fill
                        cell.font = white_text_bold
                        cell.border = thin_border
                        cell.alignment = Alignment('center')
                        column += 1

                    # go to next row, reset column
                    column = 1
                    row += 1

                    for interface, macs in interfaces.items():
                        for info in macs:
                            # write interface name
                            cell = mac_worksheet.cell(row=row, column=column)
                            cell.value = interface
                            column += 1
                            # write mac address
                            cell = mac_worksheet.cell(row=row, column=column)
                            cell.value = info['MAC']
                            column += 1
                            # write VLAN
                            cell = mac_worksheet.cell(row=row, column=column)
                            cell.value = info['VLAN']
                            column += 1
                            # write IP Address
                            cell = mac_worksheet.cell(row=row, column=column)
                            cell.value = info['IP_ADDRESS']
                            # iterate to next row and reset column
                            column = 1
                            row += 1

                # sort based on the interface
                mac_worksheet.auto_filter.add_sort_condition(f'A2:A{row - 1}')

        # save and close our file
        mac_workbook.save(output_file)
        mac_workbook.close()