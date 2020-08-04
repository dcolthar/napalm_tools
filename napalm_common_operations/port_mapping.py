from napalm import get_network_driver
import textfsm
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image



def port_mapping(network_connection):
    # command to get all interfaces switchport mode
    cmd_filter = 'inc (Switchport|Name|Administrative Mode|Native Mode VLAN|Trunking VLANs|Access Mode VLAN|Voice VLAN)'
    full_cmd = f'show interface switchport | {cmd_filter}'

    try:
        # get facts to get switch hostname
        network_facts = network_connection.get_facts()
        hostname = network_facts['hostname']

        ports_summary = {
            hostname: []
        }

        # see what interfaces are access, trunk and routed
        show_int_status = 'show interface status'
        results = network_connection.cli([show_int_status, full_cmd])
        int_status_template_file = 'textfsm_templates/show_interface_status.textfsm'
        with open(int_status_template_file) as f:
            int_status_template = textfsm.TextFSM(f)
            parsed_results = int_status_template.ParseText(results[show_int_status])
            # now that we have the info lets make a list of routed, and non routed links
            routed_links = [[i[0], 'Disabled', 'routed'] for i in parsed_results if i[3].lower() == 'routed' ]
            #switched_links = [i[0] for i in parsed_results if i[3].lower() != 'routed']

        # get textfsm going
        switched_template_file = 'textfsm_templates/show_interface_switchport.textfsm'
        with open(switched_template_file) as f:
            switched_template = textfsm.TextFSM(f)
            # parse the output
            parsed_results = switched_template.ParseText(results[full_cmd])
            #pprint(parsed_results)

        # now...join the routed links and the switched links
        full_interface_list = parsed_results + routed_links

        # AND finally add to the ports_summary dictionary
        for i in full_interface_list:
            # if we're an access port
            if i[2].lower() == 'static access':
                ports_summary[hostname].append(
                    {
                        'interface': i[0],
                        'mode': i[2],
                        'access_vlan': i[3],
                        'voice_vlan': i[5],
                        'trunk_native_vlan': 'N/A',
                        'trunk_allowed_vlans': 'N/A'
                    }
                )
            # if we're a trunk port
            elif i[2].lower() == 'trunk':
                ports_summary[hostname].append(
                    {
                        'interface': i[0],
                        'mode': i[2],
                        'access_vlan': 'N/A',
                        'voice_vlan': 'N/A',
                        'trunk_native_vlan': i[4],
                        'trunk_allowed_vlans': i[6]
                    }
                )
            # if we're a routed link
            elif i[2].lower() == 'routed':
                ports_summary[hostname].append(
                    {
                        'interface': i[0],
                        'mode': i[2],
                        'access_vlan': 'N/A',
                        'voice_vlan': 'N/A',
                        'trunk_native_vlan': 'N/A',
                        'trunk_allowed_vlans': 'N/A'
                    }
                )
            # if we're dynamic auto/desirable we're going to just grab the access vlan
            elif 'dynamic' in i[2].lower():
                ports_summary[hostname].append(
                    {
                        'interface': i[0],
                        'mode': i[2],
                        'access_vlan': i[3],
                        'voice_vlan': i[5],
                        'trunk_native_vlan': 'N/A',
                        'trunk_allowed_vlans': 'N/A'
                    }
                )

        #print(json.dumps(ports_summary, indent=1))
        return ports_summary

    except Exception as e:
        print(e)


def port_mapping_excel_creation(ports_summary, output_file, image_file):
    '''
    This will do the actual creation and writing to the excel sheet
    :param ports_summary:
    :return:
    '''

    # we'll need to offset x amount of rows and columns to fit into the existing workbook
    row_offset = 2
    column_offset = 6

    # some styling things
    bold = Font(bold=True)
    white_text_bold = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    medium_right_up_down_border = Border(left=Side(style='thin'),
                         right=Side(style='medium'),
                         top=Side(style='medium'),
                         bottom=Side(style='medium'))
    medium_left_up_down_border = Border(left=Side(style='medium'),
                         right=Side(style='thin'),
                         top=Side(style='medium'),
                         bottom=Side(style='medium'))
    medium_top_bottom_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='medium'),
                         bottom=Side(style='medium'))
    # the color you can find by getting the RGB value and then converting to hex
    grey_fill = PatternFill(
        start_color='595959',
        end_color='595959',
        fill_type='solid'
    )

    # we'll get passed a list of output we'll loop through so lets make a workbook
    # create the workbook
    ports_workbook = openpyxl.Workbook()

    #################### we can go ahead and just make the index page ahead of time here ####################
    ports_worksheet = ports_workbook.active
    # rename the sheet
    ports_worksheet.title = 'index'

    # where to start in the index sheet
    current_row = 2
    current_column = 2

    # lets set the column dimensions
    ports_worksheet.column_dimensions['A'].width = 15
    ports_worksheet.column_dimensions['B'].width = 42
    ports_worksheet.column_dimensions['C'].width = 42
    ports_worksheet.column_dimensions['D'].width = 20
    ports_worksheet.column_dimensions['E'].width = 20
    ports_worksheet.column_dimensions['F'].width = 20
    ports_worksheet.column_dimensions['G'].width = 20


    # add the company logo, first merge the cells
    ports_worksheet.merge_cells(
        start_row=current_row,
        start_column=current_column,
        end_row=current_row + 7,
        end_column=current_column + 1
    )

    ########## TODO: get formula to calculate excel cell height/width to pixels based on screen #################
    # # get height/width to make the image
    # image_height = ports_worksheet.row_dimensions[2].height
    # # image height will always be NONE if not set to a non default height, default height = 15
    # if not image_height:
    #     image_height = 15
    # image_width = ports_worksheet.column_dimensions['B'].width + ports_worksheet.column_dimensions['C'].width
    # print(f'height is {image_height}')
    #
    # # open the image with Pillow and resize then save
    # img = Image.open(image_file)
    # img = img.resize((int(image_width), int(image_height * 4)),Image.NEAREST)
    # img.save(image_file)

    # add the image to the workbook, until resize section above is complete you may need to manually resize
    image = openpyxl.drawing.image.Image(image_file)
    image.anchor = 'B3'
    ports_worksheet.add_image(image)

    # now... lets add basic info to index, need to set row to 8 but keep column at 2
    current_row = 10
    # the sections of info to write
    index_info = [
        'Project:',
        'Version:',
        'Date:',
        'Last Modified By:'
    ]

    # add info for each info section
    for i in index_info:
        # select cells, style and add info
        cell = ports_worksheet.cell(row=current_row, column=current_column)
        cell.border = medium_left_up_down_border
        cell.value = i
        cell.fill = grey_fill
        cell.font = white_text_bold

        # move to next column for blank entry
        current_column += 1
        cell = ports_worksheet.cell(row=current_row, column=current_column)
        cell.border = medium_right_up_down_border
        # reset column, increment row
        current_column -= 1
        current_row += 1

    # save the workbook
    ports_workbook.save(output_file)


    #################################### starting cell info adds ####################################
    # reset current row/columns where we need them
    home_row = 1 + row_offset
    current_row = 1 + row_offset
    home_column = 1 + column_offset
    current_column = 1 + column_offset
    # the number of cells in the switch port mapping section
    switch_data_field_width = 6

   #print(f'home column is {home_column}\ncurrent column is {current_column}\ncurrent row is {current_row}')

    # loop through outer list of items, a list of dictionaries
    for outer_value in ports_summary:
        # get key and value from each dictinary, key should be closet name, value should be switches
        for closet, switches in outer_value.items():

            # create the sheet in the workbook if it doesn't exist
            if closet not in ports_workbook.sheetnames:
                ports_workbook.create_sheet(closet)

            # set our sheet to the proper sheet
            # if we're already on the sheet no need to do so
            if ports_workbook.active.title != closet:
                ports_worksheet = ports_workbook[closet]

                # don't do the next part if the sheetname is index
                if 'index' not in ports_worksheet.title:
                    # after we create the sheet we need to do some first time things
                    # add index hyperlink, going to write by coordinate 1,1 since its always upper left cell
                    cell = ports_worksheet.cell(row=1, column=1)
                    cell.font = bold
                    cell.value = '=HYPERLINK("#index!A1", "index")'
                    cell.style = 'Hyperlink'

                    # this is the list of closet assessment info to gather
                    closet_assessment_info = [
                        'Date Collected',
                        'IDF Name / Closet Number',
                        'Management IP',
                        'Validate SSH Access',
                        'IDF Running Config Location URL',
                        'IDF Location',
                        'General Directions to IDF',
                        'Access Requirements',
                        'Physical Security, locking doors and secure environment?',
                        'Hot/Cold Aisles',
                        'Hot/Cold Aisle location relative to switches',
                        'Number and type of switches',
                        'Number of Patch Panels',
                        'Color Coded keystone jacks or patch cables for APs?',
                        'Number of racks and free rack units',
                        'Ladder Tray needed?',
                        'Horizontal and Vertical Wiremanagers',
                        'Photos of IDF (Pics of Rack, Patch Panels, Switches, UPS, Fiber Panel, etc) URL',
                        'Make and Model of UPS',
                        'Number and type of Available outlets on UPS',
                        'Number and type of Available outlets on Wall',
                        'Rack bolted down and grounded',
                        'Length of patch cables needed (if running in parallel)',
                        'Number of available Fiber Patch Panel ports and Connected Type',
                        'Number of available fiber type (OM1/OM2/OM3/OM4/OM5/OS1/OS2)',
                        'Length of fiber patch cables needed (if running in parallel',
                        'Patch Panel 1 - Rack 1',
                        'Patch Panel 2 - Rack 2',
                        'Patch Panel 3 - Rack 3',
                        'Patch Panel 4 - Rack 4'
                    ]

                    # use these rows to get us going for info writing
                    info_row_start = 2
                    info_column_start = 2
                    # write info header for Question field
                    cell = ports_worksheet.cell(row=info_row_start, column=info_column_start)
                    cell.value = 'Question'
                    cell.font = white_text_bold
                    cell.fill = grey_fill
                    cell.border = medium_left_up_down_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # write info header for Data field
                    cell = ports_worksheet.cell(row=info_row_start, column=info_column_start + 1)
                    cell.value = 'Data'
                    cell.font = white_text_bold
                    cell.fill = grey_fill
                    cell.border = medium_right_up_down_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # increment row by one to move down after writing headers
                    info_row_start += 1

                    # now lets add each of those to the list of items
                    col_width = 80
                    for item in closet_assessment_info:
                        # get column name based on index and set width
                        col_letter1 = get_column_letter(info_column_start)
                        col_letter2 = get_column_letter(info_column_start + 1)
                        ports_worksheet.column_dimensions[col_letter1].width = col_width
                        ports_worksheet.column_dimensions[col_letter2].width = col_width * 0.75
                        # format cells
                        cell = ports_worksheet.cell(row=info_row_start, column=info_column_start)
                        cell.value = item
                        cell.border = thin_border
                        # go to next cell over
                        cell = ports_worksheet.cell(row=info_row_start, column=info_column_start + 1)
                        cell.border = thin_border

                        # increment row and continue
                        info_row_start += 1



            # then we loop through each switch in the closet, interfaces is a list of dictionaries
            for switch_outer in switches:
                # used as a counter for spacing between switches
                switch_spacer = 8

                for switch, interfaces in switch_outer.items():
                    # get hostname for the sheet name
                    hostname = switch

                    # make a merged title bar
                    ports_worksheet.merge_cells(
                        start_row=current_row,
                        start_column=current_column,
                        end_row=current_row + 1,
                        end_column=current_column + switch_data_field_width
                    )
                    # add our title to the merged cell
                    cell = ports_worksheet.cell(row=current_row, column=current_column)
                    cell.value = f'Switch {hostname}'
                    cell.font = white_text_bold
                    cell.fill = grey_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                    # update current position
                    current_row += 2
                    #current_column -= switch_data_field_width

                    #print('after merged cell')
                    #print(f'home column is {home_column}\ncurrent column is {current_column}\ncurrent row is {current_row}')

                    # write titles to columns
                    col_headers = [
                        'interface', 'patch_panel_port', 'mode', 'access_vlan',
                        'voice_vlan', 'trunk_native_vlan', 'trunk_allowed_vlans'
                    ]
                    # column width values to set for each header column
                    col_width = 20
                    # iterate through column headers
                    for col in col_headers:
                        #print(f'before headers add {col}')
                        #print(f'home column is {home_column}\ncurrent column is {current_column}\ncurrent row is {current_row}')
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = col
                        cell.font = bold
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                        col_letter = get_column_letter(current_column)
                        ports_worksheet.column_dimensions[col_letter].width = col_width

                        # increment column by one
                        current_column += 1

                    # return back to start column
                    current_column -= switch_data_field_width + 1
                    current_row += 1

                    # now get the actual interfaces
                    for interface in interfaces:
                        # write interface name
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = interface['interface']
                        cell.border = thin_border
                        current_column += 1
                        # write blank for patch panel port
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.border = thin_border
                        current_column += 1
                        # write mode
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = interface['mode']
                        cell.border = thin_border
                        current_column += 1
                        # write access_vlan
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = interface['access_vlan']
                        cell.border = thin_border
                        current_column += 1
                        # write voice_vlan
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = interface['voice_vlan']
                        cell.border = thin_border
                        current_column += 1
                        # write trunk native vlan
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = interface['trunk_native_vlan']
                        cell.border = thin_border
                        current_column += 1
                        # write trunk allowed vlans
                        cell = ports_worksheet.cell(row=current_row, column=current_column)
                        cell.value = interface['trunk_allowed_vlans']
                        cell.border = thin_border
                        current_column += 1

                        # when we're done, go on row lower and go back to home column
                        current_row += 1
                        current_column -= switch_data_field_width + 1

                    # update home column just in case we'd have another switch..
                    current_column = current_column + switch_spacer
                    current_row = home_row

                    # increment switch spacer
                    switch_spacer += 8

        # reset values for next closet
        current_column = home_column
        current_row = home_row

    # Once we're complete with every closet, we need to add the hyperlinks on the main page for the closets
    ports_worksheet = ports_workbook['index']
    # remap home row/column for this
    home_row = 16
    home_column = 2
    # data fields to write
    fields = [
        'Site Name',
        'Closet Name',
        'Address',
        'Warehouse Staging',
        'Pre-Cutover Testing',
        'Cutover Complete',
        'Equipment Turned in and Logged'
    ]
    # get a list of sheet names which correlate to closets
    sheet_list = ports_workbook.sheetnames
    # build headers
    for field in fields:
        cell = ports_worksheet.cell(row=home_row, column=home_column)
        cell = ports_worksheet.cell(row=home_row, column=home_column)
        cell.value = field
        cell.fill = grey_fill
        cell.font = white_text_bold
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # conditionally add bordering
        if fields.index(field) == 0:
            cell.border = medium_left_up_down_border
        elif fields.index(field) == len(fields) - 1:
            cell.border = medium_right_up_down_border
        else:
            cell.border = medium_top_bottom_border
        # increment column
        home_column += 1

    # increment home_row and reset home_column
    # but don't set columsn all the way back as we're going to edit the Closet field
    home_row += 1
    home_column -= len(fields)

    # we're going to reuse fields here to loop through and add border

    # iterate through and do work
    for sheet_entry in sheet_list:
        if sheet_entry != 'index':
            for field in fields:
                cell = ports_worksheet.cell(row=home_row, column=home_column)
                # add a border
                cell.border = thin_border
                # if the field is the closet name we'll add it
                if field == 'Closet Name':
                    cell.font = bold
                    cell.value = f'=HYPERLINK("#{sheet_entry}!A1", "{sheet_entry}")'
                    cell.style = 'Hyperlink'
                    cell.border = thin_border
                # increment column
                home_column += 1
            # increment row and reset column
            home_row += 1
            home_column -= len(fields)

    # save and closet our file
    ports_workbook.save(output_file)
    ports_workbook.close()







